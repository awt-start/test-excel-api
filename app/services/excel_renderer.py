# app/services/excel_renderer.py
import os
import re
import logging
import time
from datetime import datetime
import openpyxl
from jinja2 import Environment, BaseLoader, TemplateSyntaxError

# 配置日志
logger = logging.getLogger(__name__)

# 正则表达式，用于匹配Jinja2的for循环的开始和结束
FOR_START_PATTERN = re.compile(r"{%\s*for\s+(\w+)\s+in\s+(\w+)\s*%}")
FOR_END_PATTERN = re.compile(r"{%\s*endfor\s*%}")

def _copy_style(source_cell, target_cell):
    """
    使用 openpyxl 内部机制，稳健地复制单元格样式。
    这是解决 'StyleArray' 问题的关键。
    """
    if source_cell.has_style:
        # 将源单元格的样式对象赋值给目标单元格
        # openpyxl 会自动处理样式的深拷贝
        target_cell._style = source_cell._style

def render_excel_template(template_path: str, context: dict) -> openpyxl.Workbook:
    """
    渲染一个包含Jinja2语法的Excel模板，支持多行循环并保留样式。
    """
    start_time = time.time()
    logger.info(f"[Excel渲染器] 开始渲染模板: {template_path}")
    logger.info(f"[Excel渲染器] 模板路径: {os.path.abspath(template_path)}")
    logger.info(f"[Excel渲染器] 渲染上下文包含字段: {list(context.keys())}")
    
    # 记录重要字段的摘要信息
    if 'notice_no' in context:
        logger.info(f"[Excel渲染器] 通知编号: {context['notice_no']}")
    if 'date' in context:
        logger.info(f"[Excel渲染器] 通知日期: {context['date']}")
    if 'projects' in context:
        logger.info(f"[Excel渲染器] 项目数量: {len(context['projects'])}")
        for i, proj in enumerate(context['projects'][:2]):  # 只记录前2个项目
            logger.info(f"[Excel渲染器] 项目{i+1}: {proj.get('project_code', 'N/A')} - {proj.get('project_name', 'N/A')}")
        if len(context['projects']) > 2:
            logger.info(f"[Excel渲染器] ... 还有{len(context['projects'])-2}个项目")
    
    try:
        wb = openpyxl.load_workbook(template_path)
        logger.info(f"[Excel渲染器] 模板文件加载成功，工作表名称: {wb.sheetnames}")
        ws = wb.active
        logger.info(f"[Excel渲染器] 激活工作表: {ws.title}")
        logger.info(f"[Excel渲染器] 工作表尺寸: {ws.max_row}行 x {ws.max_column}列")
        
    except Exception as e:
        logger.error(f"[Excel渲染器] 加载模板文件失败: {str(e)}")
        raise
    
    env = Environment(loader=BaseLoader())
    logger.info(f"[Excel渲染器] Jinja2环境初始化完成")
    
    # --- 第一步：扫描并定位所有循环块 ---
    loop_blocks = []
    max_row = ws.max_row
    max_col = ws.max_column
    
    logger.info(f"[Excel渲染器] 开始扫描循环块，扫描范围: {max_row}行 x {max_col}列")

    for row_idx in range(max_row, 0, -1):
        for col_idx in range(max_col, 0, -1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if not cell.value or not isinstance(cell.value, str):
                continue

            if FOR_END_PATTERN.search(cell.value):
                end_row, end_col = row_idx, col_idx
                start_row, start_col = None, None
                loop_var, list_name = None, None

                for r_idx in range(end_row, 0, -1):
                    for c_idx in range(max_col, 0, -1):
                        if r_idx == end_row and c_idx >= end_col:
                            continue
                        
                        start_cell = ws.cell(row=r_idx, column=c_idx)
                        if not start_cell.value or not isinstance(start_cell.value, str):
                            continue
                        
                        match = FOR_START_PATTERN.search(start_cell.value)
                        if match:
                            start_row, start_col = r_idx, c_idx
                            loop_var, list_name = match.groups()
                            break
                    if start_row:
                        break
                
                if start_row:
                    loop_blocks.append({
                        "start_row": start_row,
                        "end_row": end_row,
                        "start_col": start_col,
                        "end_col": end_col,
                        "loop_var": loop_var,
                        "list_name": list_name
                    })
                    logger.info(f"[Excel渲染器] 发现循环块 {len(loop_blocks)}: {list_name}.{loop_var} (行{start_row}-{end_row}, 列{start_col}-{end_col})")

    logger.info(f"[Excel渲染器] 扫描完成，共发现 {len(loop_blocks)} 个循环块")

    # --- 第二步：处理找到的循环块 ---
    logger.info(f"[Excel渲染器] 开始处理循环块，共{len(loop_blocks)}个")
    
    for idx, block in enumerate(loop_blocks):
        start_row = block["start_row"]
        end_row = block["end_row"]
        start_col = block["start_col"]
        end_col = block["end_col"]
        loop_var = block["loop_var"]
        list_name = block["list_name"]
        
        logger.info(f"[Excel渲染器] 处理循环块{idx+1}: {list_name}.{loop_var}")

        project_list = context.get(list_name)
        if not project_list:
            logger.warning(f"[Excel渲染器] 循环块{idx+1}: 未找到数据列表 '{list_name}'，删除模板区域")
            ws.delete_rows(start_row, (end_row - start_row + 1))
            continue
            
        logger.info(f"[Excel渲染器] 循环块{idx+1}: 找到{len(project_list)}个项目数据")

        # 提取模板区域内的所有单元格值和样式
        template_rows_data = []
        for r_idx in range(start_row, end_row + 1):
            row_data = []
            for c_idx in range(start_col, end_col + 1):
                cell = ws.cell(row=r_idx, column=c_idx)
                # 存储单元格对象本身，以便后续复制样式
                row_data.append(cell)
            template_rows_data.append(row_data)
        
        logger.info(f"[Excel渲染器] 循环块{idx+1}: 提取了{len(template_rows_data)}行模板数据")
        
        ws.delete_rows(start_row, (end_row - start_row + 1))
        logger.info(f"[Excel渲染器] 循环块{idx+1}: 删除了原始模板区域")

        # --- 第三步：为每个项目创建新行并渲染 ---
        logger.info(f"[Excel渲染器] 循环块{idx+1}: 开始为{len(project_list)}个项目创建渲染行")
        
        for i, project_item in enumerate(reversed(project_list)):
            ws.insert_rows(start_row, len(template_rows_data))
            logger.debug(f"[Excel渲染器] 循环块{idx+1}: 项目{i+1} - 插入{len(template_rows_data)}行")
            
            temp_context = {loop_var: project_item}
            project_info = project_item.get('project_code', 'N/A') if isinstance(project_item, dict) else 'N/A'
            logger.debug(f"[Excel渲染器] 循环块{idx+1}: 项目{i+1} ({project_info}) - 开始渲染单元格")
            
            rendered_count = 0
            for r_offset, template_cells in enumerate(template_rows_data):
                for c_offset, template_cell in enumerate(template_cells):
                    current_cell = ws.cell(row=start_row + r_offset, column=start_col + c_offset)
                    original_value = template_cell.value

                    # *** 修正点：直接复制整个样式对象 ***
                    if template_cell.has_style:
                        _copy_style(template_cell, current_cell)

                    # 渲染值
                    if isinstance(original_value, str) and "{{" in original_value:
                        clean_value = FOR_START_PATTERN.sub("", original_value).strip()
                        clean_value = FOR_END_PATTERN.sub("", clean_value).strip()
                        
                        if clean_value:
                            try:
                                template = env.from_string(clean_value)
                                rendered_value = template.render(temp_context)
                                current_cell.value = rendered_value
                                rendered_count += 1
                            except TemplateSyntaxError:
                                current_cell.value = original_value
                                logger.warning(f"[Excel渲染器] 循环块{idx+1}: 项目{i+1} - 模板语法错误: {original_value}")
                        else:
                            current_cell.value = None
                    else:
                        if isinstance(original_value, str) and ("{% for" in original_value or "{% endfor %}" in original_value):
                            current_cell.value = None
                        else:
                            current_cell.value = original_value
            
            logger.debug(f"[Excel渲染器] 循环块{idx+1}: 项目{i+1} - 完成渲染，渲染了{rendered_count}个变量")
        
        logger.info(f"[Excel渲染器] 循环块{idx+1}: 所有{len(project_list)}个项目渲染完成")

    # --- 第四步：处理页面中剩余的普通变量 ---
    logger.info(f"[Excel渲染器] 开始处理普通变量渲染")
    
    rendered_vars = 0
    total_cells = 0
    for row in ws.iter_rows():
        for cell in row:
            total_cells += 1
            if cell.value and isinstance(cell.value, str) and "{{" in cell.value and "{%" not in cell.value:
                try:
                    template = env.from_string(cell.value)
                    rendered_value = template.render(context)
                    cell.value = rendered_value
                    rendered_vars += 1
                    logger.debug(f"[Excel渲染器] 渲染变量: {cell.value} -> {rendered_value}")
                except TemplateSyntaxError:
                    logger.warning(f"[Excel渲染器] 普通变量模板语法错误: {cell.value}")
                    pass
    
    logger.info(f"[Excel渲染器] 普通变量渲染完成，共检查{total_cells}个单元格，渲染了{rendered_vars}个变量")
    
    elapsed_time = time.time() - start_time
    logger.info(f"[Excel渲染器] 模板渲染完成，耗时{elapsed_time:.2f}秒")
    
    return wb
                        
    return wb