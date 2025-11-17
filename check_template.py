#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import openpyxl
import os

def check_template_content(template_path, template_name):
    """检查模板文件的内容"""
    if not os.path.exists(template_path):
        print(f"❌ 模板文件不存在: {template_path}")
        return
        
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active
    
    print(f"\n=== 模板文件: {template_name} ===")
    print(f"文件路径: {template_path}")
    
    # 检查前20行内容
    content_found = False
    for row_idx, row in enumerate(ws.iter_rows(max_row=20), 1):
        row_content = []
        for col_idx, cell in enumerate(row, 1):
            if cell.value and isinstance(cell.value, str) and ('{{' in cell.value or '}}' in cell.value or '{%' in cell.value):
                row_content.append(f'[{col_idx}]{cell.value}')
                content_found = True
        
        if row_content:
            print(f"第{row_idx}行: {' | '.join(row_content)}")
    
    if not content_found:
        print("❌ 未发现任何Jinja2模板语法")
        # 显示一些原始内容
        print("\n原始内容示例:")
        for row_idx, row in enumerate(ws.iter_rows(max_row=10), 1):
            row_content = []
            for col_idx, cell in enumerate(row, 1):
                if cell.value:
                    row_content.append(f'[{col_idx}]{cell.value}')
            if row_content:
                print(f"第{row_idx}行: {' | '.join(row_content[:3])}")  # 只显示前3列
    
    wb.close()

# 检查所有模板文件
templates = [
    ('app/templates/template_横向.xlsx', '横向模板'),
    ('app/templates/template_纵向.xlsx', '纵向模板'),
    ('app/templates/template_协同创新专项.xlsx', '协同创新专项模板'),
    ('app/templates/template_协同创新专项精品.xlsx', '协同创新专项精品模板')
]

for template_path, template_name in templates:
    check_template_content(template_path, template_name)